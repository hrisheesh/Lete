from typing import List
from lete.app.schemas.section import DocumentSectionCreate
from lete.app.parsing.base import BaseParser

# Format routing:
#   .xlsx / .xlsm / .xltx / .xltm  →  openpyxl (XML-based, modern Excel)
#   .xls                            →  xlrd 2.x  (BIFF8 binary, Excel 97-2003)
#
# openpyxl CANNOT open .xls files — they raise InvalidFileException.
# xlrd 2.x CANNOT open .xlsx files — support was intentionally dropped.
# We detect at parse-time by file extension and dispatch accordingly.

OPENPYXL_EXTS = frozenset({"xlsx", "xlsm", "xltx", "xltm"})
XLRD_EXTS     = frozenset({"xls"})


class XlsxParser(BaseParser):
    """
    Parses Excel files using the correct backend per format:
      - Modern XML formats (.xlsx/.xlsm/.xltx/.xltm) → openpyxl
      - Legacy binary format (.xls, BIFF8)            → xlrd

    Each worksheet becomes its own batch of sections.
    Rows are grouped (ROWS_PER_SECTION at a time) and formatted as
    'Column: Value' pairs so each chunk is semantically self-contained
    and produces high-quality embeddings.
    """
    ROWS_PER_SECTION = 100

    # ------------------------------------------------------------------ #
    #  Public API                                                          #
    # ------------------------------------------------------------------ #

    def parse(self, file_path: str, document_id: str) -> List[DocumentSectionCreate]:
        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""

        if ext in OPENPYXL_EXTS:
            return self._parse_openpyxl(file_path, document_id)
        elif ext in XLRD_EXTS:
            return self._parse_xlrd(file_path, document_id)
        else:
            # Shouldn't happen — registry validated the ext at upload time.
            # Try openpyxl as last resort (it gives a clear error message).
            return self._parse_openpyxl(file_path, document_id)

    # ------------------------------------------------------------------ #
    #  Private helpers                                                     #
    # ------------------------------------------------------------------ #

    def _build_sections(
        self,
        sheet_name: str,
        header: List[str],
        data_rows: List[List[str]],
        document_id: str,
        section_idx_start: int,
    ) -> List[DocumentSectionCreate]:
        """Common section-building logic shared by both backends."""
        sections = []
        section_idx = section_idx_start
        header_line = ", ".join(h for h in header if h)

        for batch_start in range(0, max(len(data_rows), 1), self.ROWS_PER_SECTION):
            batch = data_rows[batch_start: batch_start + self.ROWS_PER_SECTION]
            if not batch:
                continue

            lines = [
                f"Sheet: {sheet_name}",
                f"Columns: {header_line}",
                "",
            ]
            for row in batch:
                pairs = [
                    f"{col}: {val}"
                    for col, val in zip(header, row)
                    if col and val
                ]
                if pairs:
                    lines.append(" | ".join(pairs))

            content = self._clean_text("\n".join(lines))
            if content:
                sections.append(
                    DocumentSectionCreate(
                        document_id=document_id,
                        content=content,
                        page_number=1,
                        section_index=section_idx,
                    )
                )
                section_idx += 1

        return sections

    def _parse_openpyxl(
        self, file_path: str, document_id: str
    ) -> List[DocumentSectionCreate]:
        """Handle .xlsx / .xlsm / .xltx / .xltm via openpyxl."""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sections: List[DocumentSectionCreate] = []
        section_idx = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = [str(c) if c is not None else "" for c in rows[0]]
            data_rows = [
                [str(c) if c is not None else "" for c in row]
                for row in rows[1:]
            ]
            new_sections = self._build_sections(
                sheet_name, header, data_rows, document_id, section_idx
            )
            sections.extend(new_sections)
            section_idx += len(new_sections)

        wb.close()
        return sections

    def _parse_xlrd(
        self, file_path: str, document_id: str
    ) -> List[DocumentSectionCreate]:
        """Handle legacy .xls (BIFF8, Excel 97-2003) via xlrd 2.x."""
        import xlrd

        wb = xlrd.open_workbook(file_path)
        sections: List[DocumentSectionCreate] = []
        section_idx = 0

        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            if ws.nrows == 0:
                continue

            header = [str(ws.cell_value(0, c)) for c in range(ws.ncols)]
            data_rows = [
                [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                for r in range(1, ws.nrows)
            ]
            new_sections = self._build_sections(
                sheet_name, header, data_rows, document_id, section_idx
            )
            sections.extend(new_sections)
            section_idx += len(new_sections)

        return sections
